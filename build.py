"""Wire candidates + rubric together and build the final ranked xlsx."""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from candidates import PARKS, METROS
from pop_within_radius import load_county_pop, pop_within, haversine_miles
from rubric import score_park, DG_COMMUNITY, PARK_OVERRIDES

OUT = Path(__file__).parent

counties = load_county_pop()
metro_dt = {m[0]: (m[1], m[2]) for m in METROS}

records = []
for metro, name, lat, lng, acres, notes in PARKS:
    pop, n_cty = pop_within(lat, lng, 100, counties)
    dlat, dlng = metro_dt[metro]
    dist_dt = haversine_miles(lat, lng, dlat, dlng)
    scores = score_park(metro, name, acres, pop, dist_dt)
    # Blue Sky Total: caps DG_Community penalty at 0 (positive bonus retained;
    # negative penalty zeroed out so a weak local scene doesn't torpedo an
    # otherwise iconic candidate)
    blue_sky_dg = max(0, scores["DG_Community"])
    blue_sky_total = scores["Total"] - scores["DG_Community"] + blue_sky_dg
    records.append({
        "Metro": metro,
        "Park": name,
        "Acres": acres,
        "Pop_100mi": pop,
        "Miles_to_DT": round(dist_dt, 1),
        **scores,
        "BlueSky_Total": blue_sky_total,
        "Lat": lat,
        "Lng": lng,
        "Notes": notes,
    })

df = pd.DataFrame(records).sort_values(["BlueSky_Total", "Total", "Metro"], ascending=[False, False, True]).reset_index(drop=True)
df.insert(0, "BS_Rank", range(1, len(df)+1))
df_classic = df.sort_values(["Total", "Metro"], ascending=[False, True]).reset_index(drop=True)
df_classic.insert(0, "Rank", range(1, len(df_classic)+1))
df_classic = df_classic.drop(columns=["BS_Rank"])

# ─── build workbook ──────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
H1 = Font(name="Arial", size=14, bold=True)
H2 = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color="1F4E78")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = H2
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER

def autosize(ws, max_widths=None):
    max_widths = max_widths or {}
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10,
        ) + 2
        width = min(width, max_widths.get(letter, 60))
        ws.column_dimensions[letter].width = width

# ── Sheet 1: README ─────────────────────────────────────────────────
ws = wb.create_sheet("README")
readme = [
    ("US Disc Golf Tournament Host Park Analysis", H1),
    ("Refreshed 2026 — built from the 2023 personal project", ARIAL),
    ("", None),
    ("Goal", BOLD),
    ("Identify US parks best positioned to host a USDGC-equivalent major disc golf tournament — i.e. a metro-based, large-spectator event distinct from the smaller-town majors (DDO/Emporia, etc).", ARIAL),
    ("", None),
    ("Two rankings: Blue Sky vs. Classic", BOLD),
    ("This is a BLUE SKY exercise — the goal is identifying where a major COULD work, not just where the existing DG community would already turn out. A major event in LA / NYC / Miami would PLANT the flag in those markets, not follow existing turnout.", ARIAL),
    ("  Blue Sky Total — DG Community penalty is removed (a weak local scene can't drag a park below zero on that axis). Positive DG bonuses are kept.", ARIAL),
    ("  Classic Total — original methodology with the full DG Community penalty applied.", ARIAL),
    ("", None),
    ("Rubric (each dimension -3 to +3, summed for Total)", BOLD),
    ("  Size — acres. <50: -2; 50-99: -1; 100-119: 0; 120-139: +1; 140+: +2", ARIAL),
    ("  Amenities — None / paths / +bathrooms / +buildings / +notable / +large / +giant structures (-3 to +3)", ARIAL),
    ("  Centrality — distance to downtown (straight-line). <2mi: +3; 2-5: +2; 5-10: +1; 10-20: 0; >20: -1. Manual overrides for transit quality.", ARIAL),
    ("  Parking — none → players-only → +VIP → +GA fans (-2 to +2)", ARIAL),
    ("  DG Community — metro UDisc Health (higher of Variety or Availability) mapped to -3 to +3", ARIAL),
    ("  Population — sum of population in counties whose centroid is within 100 miles of the park (NEW: replaces MSA-tier system)", ARIAL),
    ("    <1.5M: -2; 1.5-3M: -1; 3-5M: 0; 5-7M: +1; 7-10M: +2; 10M+: +3", ARIAL),
    ("  History — buildings, age, narrative hook (+) or contaminated/cemetery/etc (-)", ARIAL),
    ("  Beauty — waterfront, tree variety, vistas (+) or highway-adjacent, on a golf course, university-adjacent (-)", ARIAL),
    ("", None),
    ("Calibration anchors (from your 2023 work)", BOLD),
    ("  Winthrop / USDGC = 8 (a 'good' major)", ARIAL),
    ("  European Open (Beast) = 7", ARIAL),
    ("  DDO / Harmony Bends = negative (intentionally — small-town events; rubric is for metro majors)", ARIAL),
    ("", None),
    ("What changed from the 2023 version", BOLD),
    ("  1. Population is now computed per-PARK as 100-mile-radius population sum, using 2024 Census county estimates + 2024 county centroids. Each park gets its own population number based on its location.", ARIAL),
    ("  2. Coverage expanded: 103 candidate parks across 30 metros (was ~25 parks).", ARIAL),
    ("  3. Multiple candidates per metro so you can compare within-metro alternatives.", ARIAL),
    ("  4. Added new metros from your 'combined ≥2' tier you didn't get to score (NYC, DC, Boston, Philly, Phoenix, Seattle, Atlanta).", ARIAL),
    ("", None),
    ("Sheets in this workbook", BOLD),
    ("  Blue Sky Rankings — primary view; full table sorted by BlueSky_Total. Color scale on BlueSky_Total.", ARIAL),
    ("  Classic Rankings — your original methodology; same parks sorted by Total.", ARIAL),
    ("  By Metro — same parks grouped by metro for within-metro comparison (sorted by BlueSky_Total within metro).", ARIAL),
    ("  Criteria — the rubric reference table.", ARIAL),
    ("  Reference Anchors — Winthrop/EO/DDO from your 2023 sheet, for sanity-checking the scale.", ARIAL),
    ("", None),
    ("Known limitations", BOLD),
    ("  Amenities / History / Beauty are still judgment calls — I scored each park to the best of my knowledge but treat as starting points.", ARIAL),
    ("  Transit/centrality uses straight-line distance with manual overrides; doesn't capture last-mile transit quality precisely.", ARIAL),
    ("  Park acreage figures are from city park system pages and Wikipedia; cross-check before any production use.", ARIAL),
]
for i, (text, fnt) in enumerate(readme, 1):
    c = ws.cell(row=i, column=1, value=text)
    if fnt:
        c.font = fnt
    c.alignment = LEFT
ws.column_dimensions["A"].width = 130
for r in range(1, len(readme)+1):
    ws.row_dimensions[r].height = 22

def write_rankings_sheet(ws, frame, rank_col, sort_col_name):
    cols = [rank_col, "Metro", "Park", sort_col_name, "Total", "BlueSky_Total",
            "Size", "Amenities", "Centrality", "Parking",
            "DG_Community", "Population", "History", "Beauty",
            "Acres", "Pop_100mi", "Miles_to_DT", "Notes"]
    # remove duplicate of sort column (Total or BlueSky_Total) if it appears twice
    seen = set()
    cols = [c for c in cols if not (c in seen or seen.add(c))]
    write_header(ws, cols)
    for i, row in frame.iterrows():
        excel_row = i + 2
        for col_idx, col_name in enumerate(cols, 1):
            v = row[col_name]
            c = ws.cell(row=excel_row, column=col_idx, value=v)
            c.font = ARIAL
            c.border = BORDER
            if col_name == "Notes":
                c.alignment = LEFT
            elif col_name in ("Metro", "Park"):
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = CENTER
            if col_name == "Pop_100mi" and isinstance(v, (int, float)):
                c.number_format = "#,##0"
            if i % 2 == 1:
                c.fill = ALT_FILL
    sort_col_idx = cols.index(sort_col_name) + 1
    rule = ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B")
    ws.conditional_formatting.add(f"{get_column_letter(sort_col_idx)}2:{get_column_letter(sort_col_idx)}{len(frame)+1}", rule)
    ws.freeze_panes = "D2"
    autosize(ws, max_widths={get_column_letter(len(cols)): 70})

# ── Sheet 2: Blue Sky Rankings (primary; ignores weak-DG penalty) ──
ws = wb.create_sheet("Blue Sky Rankings")
write_rankings_sheet(ws, df, "BS_Rank", "BlueSky_Total")

# ── Sheet 3: Classic Rankings (original methodology) ───────────────
ws = wb.create_sheet("Classic Rankings")
write_rankings_sheet(ws, df_classic, "Rank", "Total")

# ── Sheet 4: By Metro ──────────────────────────────────────────────
ws = wb.create_sheet("By Metro")
df_by_metro = df.sort_values(["Metro", "BlueSky_Total"], ascending=[True, False]).reset_index(drop=True)
df_by_metro["BS_Rank"] = df_by_metro.groupby("Metro").cumcount() + 1
cols = ["BS_Rank", "Metro", "Park", "BlueSky_Total", "Total", "Size", "Amenities",
        "Centrality", "Parking", "DG_Community", "Population", "History", "Beauty",
        "Acres", "Pop_100mi", "Miles_to_DT", "Notes"]
write_header(ws, cols)
prev_metro = None
for i, row in df_by_metro.iterrows():
    excel_row = i + 2
    is_first_in_metro = row["Metro"] != prev_metro
    for col_idx, col_name in enumerate(cols, 1):
        v = row[col_name]
        c = ws.cell(row=excel_row, column=col_idx, value=v)
        c.font = ARIAL
        c.border = BORDER
        if col_name == "Notes":
            c.alignment = LEFT
        elif col_name in ("Metro", "Park"):
            c.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c.alignment = CENTER
        if col_name == "Pop_100mi" and isinstance(v, (int, float)):
            c.number_format = "#,##0"
        if is_first_in_metro:
            c.fill = PatternFill("solid", start_color="DDEBF7")
    prev_metro = row["Metro"]
bs_total_col_idx = cols.index("BlueSky_Total") + 1
rule = ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B")
ws.conditional_formatting.add(f"{get_column_letter(bs_total_col_idx)}2:{get_column_letter(bs_total_col_idx)}{len(df)+1}", rule)
ws.freeze_panes = "D2"
autosize(ws, max_widths={"Q": 70})

# ── Sheet 4: Criteria ───────────────────────────────────────────────
ws = wb.create_sheet("Criteria")
criteria_rows = [
    ["Dimension", "-3", "-2", "-1", "0", "+1", "+2", "+3", "Source / Method"],
    ["Size (acres)", "", "<50", "50-99", "100-119", "120-139", "140+", "(same)", "Acres from city park system / Wikipedia"],
    ["Amenities", "None", "Some paths", "Paths + bathrooms", "+ buildings", "+ notable structures", "+ large structures", "+ giant (stadium/amphitheatre)", "Eyeballed from park websites + Google Maps"],
    ["Centrality / Transit", "None in region", "None to park", "Some service to park", ">30 min headway", "<45 min from downtown", "<30 min from downtown", "<15 min from downtown", "Straight-line distance + manual transit override"],
    ["Parking", "", "None", "Limited street", "Enough for event+players", "+ VIP fans", "+ GA fans", "", "Eyeballed from satellite + park websites"],
    ["DG Community", "0-10", "10-20", "20-30", "30-50", "50-70", "70-90", "90-100", "UDisc Health (variety OR availability, whichever higher), mapped to -3..+3"],
    ["Population (100-mi radius)", "", "<1.5M", "1.5-3M", "3-5M", "5-7M", "7-10M", "10M+", "Sum of POPESTIMATE2024 for counties whose centroid is within 100 mi"],
    ["History", "Something horrible happened here", "Cemetery", "", "Default", "Buildings / age / narrative hook", "", "", "Wikipedia + judgment"],
    ["Beauty", "Contaminated land", "Highway adjacent", "Parking lot walk / golf course / university", "Default", "Waterfront / special trees / vistas", "Multiple positives", "Iconic", "Judgment; positives stack"],
    ["", "", "", "", "", "", "", "", ""],
    ["IMPLICIT GATES from your summary criteria", "", "", "", "", "", "", "", ""],
    ["  ~100 acres minimum, mildly wooded", "", "", "", "", "", "", "", ""],
    ["  Transit to city center <30 min", "", "", "", "", "", "", "", ""],
    ["  >300k people within 45-minute drive", "", "", "", "", "", "", "", ""],
    ["  No highways / massive sound pollution", "", "", "", "", "", "", "", ""],
    ["  No golf courses (too USDGC-similar)", "", "", "", "", "", "", "", ""],
    ["  Maybe no universities (too USDGC-similar)", "", "", "", "", "", "", "", ""],
]
for r_idx, row in enumerate(criteria_rows, 1):
    for c_idx, v in enumerate(row, 1):
        c = ws.cell(row=r_idx, column=c_idx, value=v)
        c.border = BORDER
        if r_idx == 1:
            c.font = H2
            c.fill = HDR_FILL
            c.alignment = CENTER
        elif row[0].startswith("  ") or row[0].startswith("IMPLICIT"):
            c.font = BOLD if row[0].startswith("IMPLICIT") else ARIAL
            c.alignment = LEFT
        else:
            c.font = ARIAL
            c.alignment = LEFT if c_idx in (1, 9) else CENTER
autosize(ws, max_widths={"I": 55, "A": 32})

# ── Sheet 5: Reference Anchors ─────────────────────────────────────
ws = wb.create_sheet("Reference Anchors")
anchor_cols = ["Park", "Total (your 2023)", "Notes"]
write_header(ws, anchor_cols)
anchors = [
    ("Winthrop (USDGC, Rock Hill SC)", 8, "Your major-event anchor — what a 'good' score looks like."),
    ("European Open / The Beast", 7, "Your secondary anchor — close to USDGC tier."),
    ("Forest Park, St Louis", 15, "Your highest-scoring candidate in 2023."),
    ("Grant Park, Chicago", 15, "Your highest-scoring candidate in 2023."),
    ("Denver City Park", 14, "Your top non-tied score."),
    ("National Mall, DC", 13, "High score despite DG community penalty."),
    ("Maple Hill (Leicester MA)", 4, "Established MA major, scored as reference."),
    ("Harmony Bends (Columbia MO)", -3, "Intentionally low — small-town major; rubric is for metro events."),
    ("DDO / Emporia", -4, "Same — by design."),
]
for i, (name, score, note) in enumerate(anchors, 2):
    for c_idx, v in enumerate([name, score, note], 1):
        c = ws.cell(row=i, column=c_idx, value=v)
        c.font = ARIAL
        c.border = BORDER
        c.alignment = LEFT if c_idx != 2 else CENTER
autosize(ws, max_widths={"C": 70})

out_path = OUT / "Disc Golf Tournament Parks - 2026 Refresh.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"\n{len(df)} parks across {df['Metro'].nunique()} metros")
print(f"Classic Total range: {df['Total'].min()} to {df['Total'].max()}")
print(f"Blue Sky Total range: {df['BlueSky_Total'].min()} to {df['BlueSky_Total'].max()}")
print()
print("Top 20 (Blue Sky):")
print(df.head(20)[["BS_Rank", "Metro", "Park", "BlueSky_Total", "Total", "Pop_100mi"]].to_string(index=False))
print()
print("Top 10 (Classic) for comparison:")
print(df_classic.head(10)[["Rank", "Metro", "Park", "Total", "BlueSky_Total"]].to_string(index=False))
